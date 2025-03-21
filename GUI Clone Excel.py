import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlwt

# สร้าง GUI
root = tk.Tk()
root.title("Clone Performance")
root.geometry("600x200")

# ตัวแปรเก็บ path
source_var = tk.StringVar()
dest_var = tk.StringVar()

# ฟังก์ชันเลือกไฟล์ต้นทาง
def select_source_file():
    file = filedialog.askopenfilename(
        title="เลือกไฟล์ Performance (.xlsm)",
        filetypes=[("Excel Macro-Enabled", "*.xlsm")]
    )
    if file:
        source_var.set(file)

# ฟังก์ชันเลือกไฟล์ปลายทาง
def select_dest_file():
    file = filedialog.asksaveasfilename(
        title="เลือกไฟล์ Performance (.xls)",
        defaultextension=".xls",
        filetypes=[("Excel 97-2003", "*.xls")]
    )
    if file:
        dest_var.set(file)

# ฟังก์ชันโคลน Sheet
def clone_sheet():
    source_file = source_var.get()
    dest_file = dest_var.get()

    if not source_file:
        messagebox.showerror("Error", "กรุณาเลือกไฟล์ต้นทาง!")
        return
    if not dest_file:
        messagebox.showerror("Error", "กรุณาเลือกไฟล์ปลายทาง!")
        return

    # อัปเดต GUI เพื่อแสดงว่ากำลังทำงาน
    root.config(cursor="wait")  # เปลี่ยนเคอร์เซอร์เป็น "รอ"
    tk.Label(root, text="กำลังโคลน...").grid(row=3, column=1)
    root.update()  # อัปเดตหน้าต่างทันที

    try:
        # โหลดไฟล์ต้นทาง (.xlsm)
        wb_source = load_workbook(source_file, keep_vba=True, data_only=True)  # data_only=True เพื่อให้ได้ค่าแทนสูตร
        if "SummaryByOACode" not in wb_source.sheetnames:
            raise ValueError("ไม่พบ Sheet 'SummaryByOACode' ในไฟล์ต้นทาง!")
        ws_source = wb_source["SummaryByOACode"]  # ระบุ Sheet เฉพาะ

        # สร้างไฟล์ .xls ใหม่
        wb_dest = xlwt.Workbook()
        ws_dest = wb_dest.add_sheet("SummaryByOACode")  # ชื่อ Sheet เดียวกัน

        # คัดลอกข้อมูลเป็นค่า และรักษารูปแบบพื้นฐาน
        for row_idx, row in enumerate(ws_source.rows, start=0):
            for col_idx, cell in enumerate(row, start=0):
                value = cell.value
                ws_dest.write(row_idx, col_idx, value)
                if cell.has_style:
                    style = xlwt.XFStyle()
                    # คัดลอกฟอนต์
                    if cell.font:
                        font = style.font
                        font.name = cell.font.name or "Arial"
                        font.bold = cell.font.bold or False
                        font.italic = cell.font.italic or False
                    # คัดลอกสีพื้นหลัง
                    if cell.fill and cell.fill.pattern == "solid":
                        pattern = style.pattern
                        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
                        pattern.pattern_fore_colour = 0  # สีดำ (ปรับได้)
                    # คัดลอกเส้นขอบ
                    if cell.border:
                        borders = style.borders
                        borders.left = xlwt.Borders.THIN if cell.border.left else xlwt.Borders.NO_LINE
                        borders.right = xlwt.Borders.THIN if cell.border.right else xlwt.Borders.NO_LINE
                        borders.top = xlwt.Borders.THIN if cell.border.top else xlwt.Borders.NO_LINE
                        borders.bottom = xlwt.Borders.THIN if cell.border.bottom else xlwt.Borders.NO_LINE
                    ws_dest.write(row_idx, col_idx, value, style)

        # คัดลอกความกว้างคอลัมน์
        for col in range(1, ws_source.max_column + 1):
            width = ws_source.column_dimensions[get_column_letter(col)].width
            if width:
                ws_dest.col(col - 1).width = int(width * 256)

        # บันทึกไฟล์ปลายทาง
        wb_dest.save(dest_file)
        messagebox.showinfo("สำเร็จ", f"โคลน Sheet 'SummaryByOACode' เรียบร้อยแล้วที่:\n{dest_file}")

    except Exception as e:
        messagebox.showerror("ข้อผิดพลาด", f"เกิดข้อผิดพลาด: {e}")
    finally:
        # คืนค่า GUI เป็นปกติ
        root.config(cursor="")
        tk.Label(root, text="").grid(row=3, column=1)  # ลบข้อความ "กำลังโคลน"
        root.update()

# ป้ายและช่องไฟล์ต้นทาง
tk.Label(root, text="ไฟล์ต้นทาง (.xlsm):").grid(row=0, column=0, padx=10, pady=10)
tk.Entry(root, textvariable=source_var, width=40).grid(row=0, column=1, padx=10)
tk.Button(root, text="Choose", command=select_source_file).grid(row=0, column=2, padx=10)

# ป้ายและช่องไฟล์ปลายทาง
tk.Label(root, text="ไฟล์ปลายทาง (.xls):").grid(row=1, column=0, padx=10, pady=10)
tk.Entry(root, textvariable=dest_var, width=40).grid(row=1, column=1, padx=10)
tk.Button(root, text="Choose", command=select_dest_file).grid(row=1, column=2, padx=10)

# ปุ่มเริ่มโคลน
tk.Button(root, text="Clone Excel", command=clone_sheet).grid(row=2, column=1, pady=20)

# รัน GUI
root.mainloop()